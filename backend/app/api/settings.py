"""设置 API：AI 供应商 CRUD + 健康检查，数据库参数 CRUD。"""

from __future__ import annotations

from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy.orm import Session

from app.core.deps import get_current_user
from app.db.session import get_db
from app.models.ai_provider import AIProvider as AIProviderModel
from app.models.database_param import DatabaseParam as DatabaseParamModel
from app.models.user import User
from app.schemas.settings import (
    AIProviderCreate,
    AIProviderHealthResult,
    AIProviderOut,
    AIProviderTestRequest,
    AIProviderTestResult,
    AIProviderUpdate,
    DatabaseParamCreate,
    DatabaseParamOut,
    DatabaseParamUpdate,
)

router = APIRouter()


def _require_admin(user: User) -> None:
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="仅管理员可操作设置")


# ===================== AI 供应商 =====================

@router.get("/ai-providers", response_model=list[AIProviderOut])
def list_ai_providers(
    db: Session = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    return db.query(AIProviderModel).order_by(AIProviderModel.weight.desc()).all()


@router.post("/ai-providers", response_model=AIProviderOut)
def create_ai_provider(
    body: AIProviderCreate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    _require_admin(user)
    row = AIProviderModel(
        name=body.name,
        endpoint=body.endpoint,
        api_key_enc=body.api_key,
        model=body.model,
        weight=body.weight,
        healthy=body.healthy,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.put("/ai-providers/{provider_id}", response_model=AIProviderOut)
def update_ai_provider(
    provider_id: UUID,
    body: AIProviderUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    _require_admin(user)
    row = db.query(AIProviderModel).filter(AIProviderModel.id == provider_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="AI 供应商不存在")
    data = body.model_dump(exclude_unset=True)
    if "api_key" in data:
        row.api_key_enc = data.pop("api_key")
    for k, v in data.items():
        setattr(row, k, v)
    db.commit()
    db.refresh(row)
    return row


@router.delete("/ai-providers/{provider_id}")
def delete_ai_provider(
    provider_id: UUID,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    _require_admin(user)
    row = db.query(AIProviderModel).filter(AIProviderModel.id == provider_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="AI 供应商不存在")
    db.delete(row)
    db.commit()
    return {"ok": True}


@router.post("/ai-providers/{provider_id}/health", response_model=AIProviderHealthResult)
async def check_ai_provider_health(
    provider_id: UUID,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    _require_admin(user)
    row = db.query(AIProviderModel).filter(AIProviderModel.id == provider_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="AI 供应商不存在")

    from app.ai.providers.openai_compatible import OpenAICompatibleProvider

    provider = OpenAICompatibleProvider(
        name=row.name,
        endpoint=row.endpoint,
        api_key=row.api_key_enc,
        model=row.model,
        weight=row.weight,
    )
    error = None
    try:
        ok = await provider.health()
    except Exception as exc:
        ok = False
        error = str(exc)
    row.healthy = ok
    if ok:
        from datetime import datetime, timezone
        row.last_check_at = datetime.now(timezone.utc)
    db.commit()
    return AIProviderHealthResult(id=row.id, name=row.name, healthy=ok, error=error)


@router.post("/ai-providers/test", response_model=AIProviderTestResult)
async def test_ai_provider(
    body: AIProviderTestRequest,
    _user: User = Depends(get_current_user),
):
    """用未保存的配置直接测试连通性(保存前验证,尤其针对中转站)。

    发送一条最小 chat completion 请求;成功返回 healthy=True + 延迟。
    """
    import time

    from app.ai.providers.openai_compatible import OpenAICompatibleProvider

    provider = OpenAICompatibleProvider(
        name=body.name or "test",
        endpoint=body.endpoint,
        api_key=body.api_key,
        model=body.model,
    )
    error = None
    latency_ms = None
    start = time.perf_counter()
    try:
        await provider.test_connection()
        ok = True
    except Exception as exc:  # noqa: BLE001
        ok = False
        error = str(exc)
    latency_ms = int((time.perf_counter() - start) * 1000)
    return AIProviderTestResult(healthy=ok, error=error, latency_ms=latency_ms)


# ===================== 数据库参数 =====================

@router.get("/database-params", response_model=list[DatabaseParamOut])
def list_database_params(
    category: str | None = None,
    db: Session = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    q = db.query(DatabaseParamModel)
    if category:
        q = q.filter(DatabaseParamModel.category == category)
    return q.order_by(DatabaseParamModel.category, DatabaseParamModel.model).all()


@router.post("/database-params", response_model=DatabaseParamOut)
def create_database_param(
    body: DatabaseParamCreate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    _require_admin(user)
    row = DatabaseParamModel(
        category=body.category,
        model=body.model,
        field=body.field,
        value=body.value,
        unit=body.unit,
        enabled=body.enabled,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.put("/database-params/{param_id}", response_model=DatabaseParamOut)
def update_database_param(
    param_id: UUID,
    body: DatabaseParamUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    _require_admin(user)
    row = db.query(DatabaseParamModel).filter(DatabaseParamModel.id == param_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="数据库参数不存在")
    data = body.model_dump(exclude_unset=True)
    for k, v in data.items():
        setattr(row, k, v)
    db.commit()
    db.refresh(row)
    return row


@router.delete("/database-params/{param_id}")
def delete_database_param(
    param_id: UUID,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    _require_admin(user)
    row = db.query(DatabaseParamModel).filter(DatabaseParamModel.id == param_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="数据库参数不存在")
    db.delete(row)
    db.commit()
    return {"ok": True}


@router.get("/database-params/export")
def export_database_params(
    db: Session = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    """导出全部数据库参数为 Excel 文件。"""
    import io

    from openpyxl import Workbook

    rows = db.query(DatabaseParamModel).order_by(
        DatabaseParamModel.category, DatabaseParamModel.model
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "数据库参数"
    ws.append(["分类", "型号", "字段", "值", "单位", "启用", "版本"])
    for row in rows:
        ws.append([
            row.category,
            row.model,
            row.field,
            row.value,
            row.unit or "",
            "是" if row.enabled else "否",
            row.version,
        ])
    # 设置列宽
    for col, width in zip("ABCDEFG", [15, 15, 15, 20, 10, 8, 8]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from fastapi.responses import StreamingResponse

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=database_params.xlsx"},
    )


@router.post("/database-params/import")
def import_database_params(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """从 Excel 文件导入数据库参数，覆盖式更新（同名同型号同字段先删后建）。"""
    _require_admin(user)
    import io

    from openpyxl import load_workbook

    content = file.file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    # 读取表头，建立列索引
    headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {}
    for idx, h in enumerate(headers):
        col_map[h] = idx

    # 允许的表头名
    required = ["分类", "型号", "字段", "值"]
    for r in required:
        if r not in col_map:
            raise HTTPException(status_code=400, detail=f"Excel 缺少必需列: {r}")

    imported = 0
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[col_map["分类"]]:
            skipped += 1
            continue
        category = str(row[col_map["分类"]]).strip()
        model = str(row[col_map["型号"]]).strip() if "型号" in col_map and row[col_map["型号"]] else ""
        field = str(row[col_map["字段"]]).strip() if "字段" in col_map and row[col_map["字段"]] else ""
        value = str(row[col_map["值"]]).strip() if "值" in col_map and row[col_map["值"]] else ""

        unit = ""
        if "单位" in col_map and col_map["单位"] < len(row) and row[col_map["单位"]]:
            unit = str(row[col_map["单位"]]).strip()

        enabled = True
        if "启用" in col_map and col_map["启用"] < len(row) and row[col_map["启用"]]:
            enabled_str = str(row[col_map["启用"]]).strip()
            enabled = enabled_str in ("是", "true", "True", "1", "启用")

        version = 1
        if "版本" in col_map and col_map["版本"] < len(row) and row[col_map["版本"]]:
            try:
                version = int(row[col_map["版本"]])
            except (TypeError, ValueError):
                version = 1

        if not all([category, model, field, value]):
            skipped += 1
            continue

        # 覆盖式：先删同键
        db.query(DatabaseParamModel).filter(
            DatabaseParamModel.category == category,
            DatabaseParamModel.model == model,
            DatabaseParamModel.field == field,
            DatabaseParamModel.version == version,
        ).delete()

        row_obj = DatabaseParamModel(
            category=category,
            model=model,
            field=field,
            value=value,
            unit=unit or None,
            enabled=enabled,
            version=version,
        )
        db.add(row_obj)
        imported += 1

    db.commit()
    return {"imported": imported, "skipped": skipped}
