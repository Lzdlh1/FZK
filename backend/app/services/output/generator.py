"""最终输出 Excel 生成。

负责:
- 把审核后的字段值回填进 Univer 快照(``fill_snapshot_with_values``);
- 把填充后的快照转成 xlsx 字节流(``snapshot_to_xlsx_bytes``,后端独立用 openpyxl,
  不依赖前端 SheetJS;ADR#13 仅保留值,不写 Excel 公式);
- 输出文件名构造(``build_output_filename`` / ``sanitize_filename``)。

Univer 快照结构(对应前端 IUniverSnapshot / IWorkbookData):
    {
      "id": "workbook-1",
      "sheetOrder": ["s1", ...],
      "sheets": {
        "s1": {"id": "s1", "name": "Sheet1", "cellData": {rowIndex: {colIndex: {"v": value}}}},
        ...
      }
    }
行/列号从 0 开始;JSONB 反序列化后键为字符串,本模块读取时统一 ``int()`` 处理。
"""

from __future__ import annotations

import io
import os
import re
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_ILLEGAL_FILENAME_CHARS = set('\\/:*?"<>|')
_A1_RE = re.compile(r"^([A-Za-z]+)(\d+)$")

# Univer 边框样式号 -> openpyxl 边框样式字符串
_BORDER_STYLE_MAP = {
    1: "thin", 2: "hair", 3: "dotted", 4: "dashed",
    5: "dashDot", 6: "dashDotDot", 7: "double", 8: "medium",
    9: "mediumDashed", 10: "mediumDashDot", 11: "mediumDashDotDot",
    12: "slantDashDot", 13: "thick",
}
# Univer 水平/垂直对齐数字 -> openpyxl 字符串
_HT_MAP = {1: "left", 2: "center", 3: "right", 4: "justify"}
_VT_MAP = {1: "top", 2: "center", 3: "bottom"}


def sanitize_filename(name: str) -> str:
    """过滤文件名非法字符 \\/:*?"<>| 替换为 _。空名回退为“输出”。"""
    if not name:
        return "输出"
    return "".join("_" if ch in _ILLEGAL_FILENAME_CHARS else ch for ch in name)


def build_output_filename(drawing_name: str, custom_name: str | None) -> str:
    """构造输出文件名:{基础名}_{YYYYMMDDHHMM}.xlsx(UTC,到分钟)。

    基础名:用户传入 custom_name,或“图纸名(去扩展名)+工艺辅助卡”。
    """
    if custom_name:
        base = custom_name
    else:
        stem = os.path.splitext(drawing_name)[0] if drawing_name else "输出"
        base = f"{stem}工艺辅助卡"
    base = sanitize_filename(base)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d%H%M")
    return f"{base}_{ts}.xlsx"


def a1_to_row_col(a1: str) -> tuple[int, int] | None:
    """A1 引用(如 "AB12")转 {row, col}(0-based);失败返回 None。"""
    if not a1:
        return None
    m = _A1_RE.match(a1.strip())
    if not m:
        return None
    letters = m.group(1).upper()
    col = 0
    for ch in letters:
        col = col * 26 + (ord(ch) - 64)
    col -= 1
    row = int(m.group(2)) - 1
    return row, col


def _get(obj: Any, key: str, default: Any = None) -> Any:
    """兼容 ORM 对象(属性)与 dict(键)读取。"""
    if isinstance(obj, dict):
        return obj.get(key, default)
    return getattr(obj, key, default)


def _get_row(cell_data: dict, row: int) -> dict:
    """按行号取行 dict,兼容 int/str 键(JSONB 反序列化后为 str)。"""
    return cell_data.get(row) or cell_data.get(str(row)) or {}


def fill_snapshot_with_values(
    univer_snapshot: dict, variables: list, fields: dict
) -> dict:
    """把 fields[var.name] 的值写入快照对应 sheet+cell。

    - 仅处理 enabled 变量(缺省视为 True);
    - fields[var.name] 可为 {value: ...} dict 或裸值;value 为 None 则留空(跳过);
    - 通过变量 sheet 名称匹配 snapshot.sheets[*].name,cell 用 A1 定位;
    - 合并单元格的左上锚点由变量 cell 直接指定,无需特殊处理;
    - 返回填充后的新 snapshot(浅拷贝,不修改入参的 sheets 结构)。
    """
    snapshot = dict(univer_snapshot or {})
    sheets = dict(snapshot.get("sheets") or {})

    # 名称 -> sheetId 映射(同名取第一个)
    name_to_id: dict[str, str] = {}
    for sid, sheet in sheets.items():
        if isinstance(sheet, dict):
            sname = sheet.get("name", sid)
            name_to_id.setdefault(sname, sid)

    for var in variables or []:
        if _get(var, "enabled", True) is False:
            continue
        name = _get(var, "name")
        sheet_name = _get(var, "sheet")
        cell_ref = _get(var, "cell")
        if not name or not sheet_name or not cell_ref:
            continue

        fdata = (fields or {}).get(name)
        if fdata is None:
            continue
        value = fdata.get("value") if isinstance(fdata, dict) else fdata
        if value is None:
            continue

        rc = a1_to_row_col(cell_ref)
        if rc is None:
            continue
        row, col = rc

        sid = name_to_id.get(sheet_name) or (
            sheet_name if sheet_name in sheets else None
        )
        if sid is None:
            continue

        sheet = dict(sheets.get(sid) or {})
        cell_data = dict(sheet.get("cellData") or {})
        rrow = dict(_get_row(cell_data, row))
        # 保留原 cell 的样式(s)/富文本结构,只更新 v,避免清掉模板样式
        existing = rrow.get(col)
        if isinstance(existing, dict):
            new_cell = dict(existing)
            new_cell["v"] = value
        else:
            new_cell = {"v": value}
        rrow[col] = new_cell
        cell_data[row] = rrow
        sheet["cellData"] = cell_data
        sheets[sid] = sheet

    snapshot["sheets"] = sheets
    return snapshot


def _coerce_cell_value(v: Any) -> Any:
    """openpyxl 仅接受 str/int/float/bool/datetime;其余转字符串。"""
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        return v
    return str(v)


def _color_to_argb(color: Any) -> str | None:
    """Univer 颜色 {rgb:'#RRGGBB'} -> openpyxl 'FFRRGGBB'。"""
    if isinstance(color, dict):
        rgb = color.get("rgb")
        if isinstance(rgb, str):
            c = rgb.lstrip("#")
            if len(c) == 6:
                return "FF" + c.upper()
    return None


def _extract_text_from_p(p: Any) -> str | None:
    """从 Univer 富文本对象 p 提取纯文本。

    p.body.dataStream 是完整文本(末尾可能带 \\r/\\n 分隔符)。
    """
    if not isinstance(p, dict):
        return None
    body = p.get("body")
    if isinstance(body, dict):
        ds = body.get("dataStream")
        if isinstance(ds, str) and ds:
            # 去掉末尾的 \r\n / \r / \n
            return ds.rstrip("\r\n")
    return None


def _resolve_cell_style(cell: dict, styles: dict) -> dict | None:
    """解析 cell.s 为样式对象。s 可以是 styles 表的 key(str)或内联样式 dict。"""
    s = cell.get("s") if isinstance(cell, dict) else None
    if s is None:
        return None
    if isinstance(s, str):
        # 引用 styles 表
        return styles.get(s) if isinstance(styles, dict) else None
    if isinstance(s, dict):
        return s
    return None


def _apply_style(ws_cell: Any, style: dict) -> None:
    """把 Univer 样式 dict 映射到 openpyxl 单元格。

    支持字体(ff/fs/bl/it/ul/st/cl)、填充(bg)、边框(bd)、
    对齐(ht/vt/tb/tr)、数字格式(n)。
    """
    if not isinstance(style, dict):
        return
    # 字体
    font_kw: dict[str, Any] = {}
    if "ff" in style:
        font_kw["name"] = style["ff"]
    if "fs" in style:
        try:
            font_kw["size"] = int(style["fs"])
        except (TypeError, ValueError):
            pass
    if style.get("bl"):
        font_kw["bold"] = True
    if style.get("it"):
        font_kw["italic"] = True
    if style.get("ul"):
        font_kw["underline"] = "single"
    if style.get("st"):
        font_kw["strike"] = True
    cl = _color_to_argb(style.get("cl"))
    if cl:
        font_kw["color"] = cl
    if font_kw:
        ws_cell.font = Font(**font_kw)
    # 填充
    bg = _color_to_argb(style.get("bg"))
    if bg:
        ws_cell.fill = PatternFill(fill_type="solid", fgColor=bg)
    # 边框
    bd = style.get("bd")
    if isinstance(bd, dict) and bd:
        side_kw: dict[str, Side | None] = {"top": None, "bottom": None, "left": None, "right": None}
        for src, dst in (("t", "top"), ("b", "bottom"), ("l", "left"), ("r", "right")):
            edge = bd.get(src)
            if isinstance(edge, dict):
                sty_num = edge.get("s")
                sty_str = _BORDER_STYLE_MAP.get(sty_num) if isinstance(sty_num, int) else None
                if sty_str:
                    ec = _color_to_argb(edge.get("cl")) or "FF000000"
                    side_kw[dst] = Side(style=sty_str, color=ec)
        if any(side_kw.values()):
            ws_cell.border = Border(**side_kw)
    # 对齐
    align_kw: dict[str, Any] = {}
    if style.get("ht") in _HT_MAP:
        align_kw["horizontal"] = _HT_MAP[style["ht"]]
    if style.get("vt") in _VT_MAP:
        align_kw["vertical"] = _VT_MAP[style["vt"]]
    if style.get("tb") == 3:
        align_kw["wrap_text"] = True
    tr = style.get("tr")
    if isinstance(tr, dict) and isinstance(tr.get("a"), (int, float)):
        align_kw["text_rotation"] = int(tr["a"])
    if align_kw:
        ws_cell.alignment = Alignment(**align_kw)
    # 数字格式
    n = style.get("n")
    if isinstance(n, dict):
        pattern = n.get("pattern")
        if isinstance(pattern, str) and pattern:
            ws_cell.number_format = pattern


def snapshot_to_xlsx_bytes(snapshot: dict) -> bytes:
    """把 Univer 快照转成 xlsx 字节流(还原模板样式/合并单元格/列宽行高)。

    按 sheetOrder(缺失则 sheets 插入序)逐 sheet 创建 openpyxl 工作表,
    遍历 cellData 写入值并应用样式;还原 mergeData/columnData/rowData。
    富文本(p)在没有 v 时作为纯文本 fallback,避免模板静态标签丢失。
    """
    snapshot = snapshot or {}
    sheets = snapshot.get("sheets") or {}
    styles = snapshot.get("styles") or {}
    order = snapshot.get("sheetOrder")
    if order:
        sheet_ids = [sid for sid in order if sid in sheets]
    else:
        sheet_ids = list(sheets.keys())

    wb = Workbook()
    default_ws = wb.active
    first = True
    for sid in sheet_ids:
        sheet = sheets.get(sid) or {}
        if not isinstance(sheet, dict):
            continue
        name = sheet.get("name") or sid
        if first:
            ws = default_ws
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(title=name)

        # 列宽(Univer 像素 -> openpyxl 字符宽度近似:px ≈ w*7+5)
        column_data = sheet.get("columnData") or {}
        if isinstance(column_data, dict):
            for c_key, cinfo in column_data.items():
                try:
                    c = int(c_key)
                except (ValueError, TypeError):
                    continue
                if isinstance(cinfo, dict) and isinstance(cinfo.get("w"), (int, float)):
                    ws.column_dimensions[get_column_letter(c + 1)].width = round(
                        (cinfo["w"] - 5) / 7, 2
                    )
        # 行高(Univer 像素 -> openpyxl 点:pt ≈ px*3/4)
        row_data = sheet.get("rowData") or {}
        if isinstance(row_data, dict):
            for r_key, rinfo in row_data.items():
                try:
                    r = int(r_key)
                except (ValueError, TypeError):
                    continue
                if isinstance(rinfo, dict) and isinstance(rinfo.get("h"), (int, float)):
                    ws.row_dimensions[r + 1].height = round(rinfo["h"] * 3 / 4, 2)

        # 合并单元格
        merge_data = sheet.get("mergeData") or []
        if isinstance(merge_data, list):
            for m in merge_data:
                if not isinstance(m, dict):
                    continue
                try:
                    sr = int(m["startRow"]) + 1
                    er = int(m["endRow"]) + 1
                    sc = int(m["startColumn"]) + 1
                    ec = int(m["endColumn"]) + 1
                except (KeyError, TypeError, ValueError):
                    continue
                try:
                    ws.merge_cells(
                        start_row=sr, end_row=er, start_column=sc, end_column=ec
                    )
                except Exception:  # noqa: BLE001 - 合并冲突忽略
                    pass

        # 单元格值 + 样式
        cell_data = sheet.get("cellData") or {}
        if not isinstance(cell_data, dict):
            continue
        for r_key, row in cell_data.items():
            try:
                r = int(r_key)
            except (ValueError, TypeError):
                continue
            if not isinstance(row, dict):
                continue
            for c_key, cell in row.items():
                try:
                    c = int(c_key)
                except (ValueError, TypeError):
                    continue
                if not isinstance(cell, dict):
                    # 裸值
                    value = _coerce_cell_value(cell)
                    if value is not None:
                        ws.cell(row=r + 1, column=c + 1, value=value)
                    continue
                # 优先用 v;v 为空时从富文本 p 提取纯文本(模板静态标签常存于 p)
                value = cell.get("v")
                if value is None:
                    value = _extract_text_from_p(cell.get("p"))
                value = _coerce_cell_value(value)
                if value is None:
                    # 即使无值也应用样式(空但有样式的单元格需保留边框/底纹)
                    ws_cell = ws.cell(row=r + 1, column=c + 1)
                else:
                    ws_cell = ws.cell(row=r + 1, column=c + 1, value=value)
                # 应用样式
                style = _resolve_cell_style(cell, styles)
                if style:
                    _apply_style(ws_cell, style)

    if first:
        # 没有任何 sheet:保留默认空工作表,重命名为 Sheet1
        default_ws.title = "Sheet1"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
