"""output generator 单测:文件名/过滤/快照填充/xlsx 转换。"""

import io

from openpyxl import load_workbook

from app.services.output.generator import (
    a1_to_row_col,
    build_output_filename,
    fill_snapshot_with_values,
    sanitize_filename,
    snapshot_to_xlsx_bytes,
)


def test_sanitize_filename():
    assert sanitize_filename("a/b:c*d?e\"f<g>h|i") == "a_b_c_d_e_f_g_h_i"
    assert sanitize_filename("正常名称") == "正常名称"
    assert sanitize_filename("") == "输出"


def test_build_output_filename():
    name = build_output_filename("线束A.pdf", None)
    assert name.endswith(".xlsx")
    assert "线束A工艺辅助卡_" in name
    for ch in '\\/:*?"<>|':
        assert ch not in name

    custom = build_output_filename("ignored.pdf", "我的卡片")
    assert custom.startswith("我的卡片_")
    assert custom.endswith(".xlsx")

    empty = build_output_filename("", None)
    assert empty.startswith("输出工艺辅助卡_") and empty.endswith(".xlsx")


def test_a1_to_row_col():
    assert a1_to_row_col("A1") == (0, 0)
    assert a1_to_row_col("B2") == (1, 1)
    assert a1_to_row_col("AA10") == (9, 26)
    assert a1_to_row_col("") is None
    assert a1_to_row_col("not-a-cell") is None


def _make_snapshot():
    return {
        "id": "workbook-1",
        "sheetOrder": ["s1"],
        "sheets": {
            "s1": {"id": "s1", "name": "Sheet1", "cellData": {}},
        },
    }


def test_fill_snapshot_with_values():
    snap = _make_snapshot()
    variables = [
        {"name": "varA", "sheet": "Sheet1", "cell": "A1", "enabled": True},
        {"name": "varB", "sheet": "Sheet1", "cell": "B2", "enabled": True},
    ]
    fields = {
        "varA": {"value": 100, "confidence": 0.9},
        "varB": {"value": "DJ7021"},
    }
    filled = fill_snapshot_with_values(snap, variables, fields)
    cd = filled["sheets"]["s1"]["cellData"]
    # A1 -> row 0, col 0
    assert cd[0][0]["v"] == 100
    # B2 -> row 1, col 1
    assert cd[1][1]["v"] == "DJ7021"
    # 原始入参未被修改
    assert snap["sheets"]["s1"]["cellData"] == {}


def test_fill_skips_disabled_and_null():
    snap = _make_snapshot()
    variables = [
        {"name": "ok", "sheet": "Sheet1", "cell": "A1", "enabled": True},
        {"name": "off", "sheet": "Sheet1", "cell": "A2", "enabled": False},
        {"name": "nullv", "sheet": "Sheet1", "cell": "A3", "enabled": True},
        {"name": "missing", "sheet": "Sheet1", "cell": "A4", "enabled": True},
    ]
    fields = {"ok": {"value": 1}, "off": {"value": 2}, "nullv": {"value": None}}
    filled = fill_snapshot_with_values(snap, variables, fields)
    cd = filled["sheets"]["s1"]["cellData"]
    assert cd[0][0]["v"] == 1
    assert 1 not in cd  # A2 disabled
    assert 2 not in cd  # A3 null value
    assert 3 not in cd  # A4 missing field


def test_snapshot_to_xlsx_bytes():
    snap = _make_snapshot()
    variables = [
        {"name": "varA", "sheet": "Sheet1", "cell": "A1"},
        {"name": "varB", "sheet": "Sheet1", "cell": "B2"},
    ]
    fields = {"varA": {"value": 100}, "varB": {"value": "DJ7021"}}
    filled = fill_snapshot_with_values(snap, variables, fields)

    data = snapshot_to_xlsx_bytes(filled)
    assert isinstance(data, bytes) and len(data) > 0

    wb = load_workbook(io.BytesIO(data))
    assert "Sheet1" in wb.sheetnames
    ws = wb["Sheet1"]
    assert ws.cell(row=1, column=1).value == 100
    assert ws.cell(row=2, column=2).value == "DJ7021"


def test_snapshot_to_xlsx_bytes_string_keys():
    """JSONB 反序列化后 cellData 键为字符串,验证也能正确转换。"""
    snap = {
        "id": "wb",
        "sheetOrder": ["s1"],
        "sheets": {
            "s1": {
                "id": "s1",
                "name": "Sheet1",
                "cellData": {"0": {"0": {"v": "静态值"}}},
            }
        },
    }
    data = snapshot_to_xlsx_bytes(snap)
    wb = load_workbook(io.BytesIO(data))
    assert wb["Sheet1"].cell(row=1, column=1).value == "静态值"
